import openpyxl
import random
from openpyxl.styles import Font,Color,PatternFill
from tkinter import *
from tkinter import messagebox
from tkinter import filedialog

root = Tk()
root.geometry("600x600")
#root.resizable(False,False)
#frame=LabelFrame(root,padx=200,pady=20,bg='orange')
#frame.pack()
label1=Label(root,text="SPORTS APP",font=(20),fg='white',bg='red',padx=200,pady=15)
label1.pack(pady=10)


path = filedialog.askopenfilename(title = 'Select an Excel file')
print(path)

def getTeamList():
    ##Load the required workbook
    wb = openpyxl.load_workbook(path)
    #sheet = wb.active
    ws=wb['Sheet1']
    ##sheet = wb.sheetnames
    ##print(wb.active.title)
    dataList = []
    value_range=ws['A2':'B20']

    for a,b in value_range:
        dataList.append(a.value)
    #return dataList

    ## Checking which colmn has team names
    #cols = sheet.max_column + 1
    #for y in  range(1, cols):
        #header = sheet.cell(1, y).value
        #print(header)
        #if header.upper() == 'TEAM NAMES':
            #pos = y
    #print(pos)
    #Generating the list of team names
    #rows = sheet.max_row + 1
    #for x in range(2, rows):
        #dataList.append(sheet.cell(x, pos).value)
    return dataList

#################################### generating new team match up ###################################################

#check the no of element in list is even or odd
def getMatch(teamsList, n):
    computedTeamList = []
    a = n%2
    if a==1:
        #b = str(input("Number of teams is odd.So please enter priority team name:"))
        b=random.choice(teamsList)
        teamsList.remove(b)
    #shuffle
    random.shuffle(teamsList)
    #print(Teams)
    length=len(teamsList)
    i =0
    #output
    while i < length:
     #   print("MATCH GENERATED ARE:")
        #print("Team", teamsList[i], "VS", "Team", teamsList[i + 1])
        outText = "{0} VS {1}".format(teamsList[i], teamsList[i+1])
        computedTeamList.append(outText)
        i+=2

    ## adding  the one merit team at the end of list
    computedTeamList.append(b+" (Priority Team)")
    return computedTeamList

############################################## write to Exel ###############################3
def writeToExel(newTeamList):
    wb = openpyxl.load_workbook(path)
   # ws=wb['Output Sheet']
    #wb.remove(ws)
    #wb.save(path)
    ##load workbook
    newSheet = wb.create_sheet()                 ##create new sheet
    newSheet.title = "Output Sheet"
    x =1                ##x=1
    newSheet.cell(row=x, column=1, value="GENERATED MATCH LIST")   ## first row = header
    x +=1               ## x=2
    for i in range(len(newTeamList)):
        newSheet.cell(row=x, column= 1, value=newTeamList[i])
        x +=1
    wb.save(path)

########################################## Adding styles ########################################
def add_style(newTeamList):
    wb = openpyxl.load_workbook(path)
    ws=wb.active
    heading=Font(size=20,color="FF0000")
    a1=ws['A1']
    a1.font=heading
    style=Font(size=12,color="FFA500")
    for i in range(2,len(newTeamList)+2):
        ws.cell(row=i,column=1).font=style
    color=PatternFill(patternType='solid',fgColor='87CEEB')
    ws['A1'].fill=color
    wb.save(path)

############################ Adjusting rows and columns ###############################
def adjust_row_column():
    wb = openpyxl.load_workbook(path) 
    ws=wb['Output Sheet']
    ws.column_dimensions['A'].width=40
    wb.save(path)


teamsList = []
newTeamList = []
## team list from exel sheet
teamsList = getTeamList()
print(teamsList)
n = len(teamsList)
#print(n)
## new team list of matches from getmatch()
newTeamList = getMatch(teamsList, n)
print(newTeamList)
## new exel sheet of team matches
writeToExel(newTeamList)
add_style(newTeamList)
adjust_row_column()

wb = openpyxl.load_workbook(path)
ws=wb['Output Sheet']

column = ws['A']
#def disp():
    #displayList = ''
    #for a in column:
        #displayList = f'{displayList + str(a.value)}\n'
    #label3.config(text=displayList)

def disp():
    scrollbar1=Scrollbar(frame,orient=VERTICAL)
    scrollbar1.pack(side=RIGHT,fill=Y)
    scrollbar2=Scrollbar(frame,orient=HORIZONTAL)
    scrollbar2.pack(side=BOTTOM,fill=X)
    listbox=Listbox(frame,bg='orange',yscrollcommand=scrollbar1.set,xscrollcommand=scrollbar2.set)
    for a in column:
        listbox.insert(END, a.value)
        listbox.pack(fill=BOTH)
    scrollbar1.config(command=listbox.yview)
    scrollbar2.config(command=listbox.xview)
def exit():
    message = messagebox.askquestion("Confirmation for Exiting", "Are you sure you want to exit?")
    if message == 'yes':
        root.quit()

label2=Label(root,text="Excel file was selected successfully.",font=(5),fg='green',bg='light green',padx=100)
label2.pack(pady=10)
label3=Label(root,text='Please click "Get Match List" button to get match list.',font=(5))
label3.pack(pady=10)
button1=Button(root,text='Get Match List',command=disp,font=(5),bg='blue',padx=10)
button1.pack(pady=10)
#print(button1)
frame=LabelFrame(root)
frame.pack(pady=50)
#label4=Label(frame,text="",font=(100),pady=20)
#label4.pack()
button3=Button(root,text="Exit",command=exit,font=(5),bg='grey',padx=10)
button3.pack(side=BOTTOM,pady=10)



root.mainloop()